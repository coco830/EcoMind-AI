from __future__ import annotations

"""Regulator aggregation service (T+1, aggregate-only)."""

import io
import json
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Any
from zoneinfo import ZoneInfo

import h3
import structlog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.device import Device, DeviceStatus
from app.models.monitoring_mysql import MonitoringDailyStats
from app.models.organization import Organization, OrganizationType
from app.models.self_inspection import SelfInspectionReport, SelfInspectionData
from app.models.user import User

logger = structlog.get_logger(__name__)

RISK_LEVELS = [
    (80, "L5"),
    (60, "L4"),
    (40, "L3"),
    (20, "L2"),
    (0, "L1"),
]

DEFAULT_MIN_SAMPLE = 10
MIN_SAMPLE_BY_INDUSTRY: dict[str, int] = {
    "steel": 8,
    "petrochemical": 6,
    "chemical": 6,
    "cement": 6,
    "coking": 6,
    "nonferrous_metal": 6,
    "glass_ceramic": 6,
    "paper_making": 6,
    "textile_dyeing": 6,
    "electroplating": 6,
}


@dataclass
class RegulatorScope:
    level: str | None
    codes: list[str]


def _resolve_target_date(target_date: date | None) -> date:
    now = datetime.now(ZoneInfo("Asia/Shanghai")).date()
    latest = now - timedelta(days=1)
    if target_date is None or target_date > latest:
        return latest
    return target_date


def _resolve_date_range(
    start_date: date | None,
    end_date: date | None,
    default_days: int = 30,
) -> tuple[date, date]:
    end = _resolve_target_date(end_date)
    start = start_date or (end - timedelta(days=default_days - 1))
    if start > end:
        start = end
    return start, end


def _resolve_month_range(year: int, month: int) -> tuple[date, date]:
    start = date(year, month, 1)
    if month == 12:
        end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(year, month + 1, 1) - timedelta(days=1)
    return start, end


def _safe_divide(numerator: float, denominator: float) -> float:
    if denominator <= 0:
        return 0.0
    return numerator / denominator


def _risk_level(score: float) -> str:
    for threshold, level in RISK_LEVELS:
        if score >= threshold:
            return level
    return "L1"


def _parse_json_list(value: str | None) -> list[str]:
    if not value:
        return []
    try:
        data = json.loads(value)
    except json.JSONDecodeError:
        return []
    if not isinstance(data, list):
        return []
    return [str(item) for item in data]


def _parse_numeric_value(raw: str) -> float | None:
    if not raw:
        return None
    cleaned = raw.strip().replace(" ", "")
    cleaned = cleaned.replace("×", "x").replace("X", "x")
    if "x10^" in cleaned:
        base, exp = cleaned.split("x10^", 1)
        try:
            return float(base) * (10 ** int(exp))
        except ValueError:
            return None
    if "x10" in cleaned:
        base, exp = cleaned.split("x10", 1)
        try:
            return float(base) * (10 ** int(exp))
        except ValueError:
            return None
    try:
        return float(cleaned)
    except ValueError:
        return None


class RegulatoryService:
    """Aggregate regulator data with strict masking."""

    def __init__(self, db: AsyncSession):
        self.db = db

    def _get_scope(self, user: User) -> RegulatorScope | None:
        if user.is_superadmin:
            return None
        org = getattr(user, "organization", None)
        if not org or getattr(org, "org_type", None) != OrganizationType.REGULATOR.value:
            return RegulatorScope(level=None, codes=[])
        return RegulatorScope(
            level=org.jurisdiction_level,
            codes=_parse_json_list(org.jurisdiction_codes),
        )

    def _apply_org_scope(
        self,
        query,
        scope: RegulatorScope | None,
        region_code: str | None,
        park_code: str | None,
    ):
        if scope and scope.codes:
            if scope.level == "park":
                query = query.where(Organization.park_code.in_(scope.codes))
            else:
                query = query.where(Organization.region_code.in_(scope.codes))
        if region_code:
            query = query.where(Organization.region_code == region_code)
        if park_code:
            query = query.where(Organization.park_code == park_code)
        return query

    async def _fetch_enterprise_orgs(
        self,
        user: User,
        region_code: str | None,
        park_code: str | None,
    ) -> list[Organization]:
        scope = self._get_scope(user)
        if scope and not scope.codes:
            return []
        query = select(Organization).where(
            Organization.org_type == OrganizationType.ENTERPRISE.value
        )
        query = self._apply_org_scope(query, scope, region_code, park_code)
        result = await self.db.execute(query)
        return result.scalars().all()

    async def _device_stats_by_org(self, org_ids: list[str]) -> dict[str, dict[str, int]]:
        if not org_ids:
            return {}
        query = (
            select(
                Device.org_id,
                func.count(Device.id).label("total"),
                func.count().filter(Device.status == DeviceStatus.ONLINE.value).label("online"),
                func.count().filter(Device.status == DeviceStatus.OFFLINE.value).label("offline"),
                func.count().filter(Device.status == DeviceStatus.ALARM.value).label("alarm"),
            )
            .where(Device.org_id.in_(org_ids))
            .group_by(Device.org_id)
        )
        result = await self.db.execute(query)
        rows = result.all()
        stats: dict[str, dict[str, int]] = {}
        for row in rows:
            stats[str(row.org_id)] = {
                "total": int(row.total or 0),
                "online": int(row.online or 0),
                "offline": int(row.offline or 0),
                "alarm": int(row.alarm or 0),
            }
        return stats

    async def _daily_stats_by_org(
        self,
        org_ids: list[str],
        target_date: date,
    ) -> dict[str, dict[str, float]]:
        if not org_ids:
            return {}
        query = (
            select(
                MonitoringDailyStats.org_id,
                func.sum(MonitoringDailyStats.data_count).label("data_count"),
                func.sum(MonitoringDailyStats.exceed_count).label("exceed_count"),
                func.sum(MonitoringDailyStats.invalid_count).label("invalid_count"),
            )
            .where(
                and_(
                    MonitoringDailyStats.stat_date == target_date,
                    MonitoringDailyStats.org_id.in_(org_ids),
                )
            )
            .group_by(MonitoringDailyStats.org_id)
        )
        result = await self.db.execute(query)
        rows = result.all()
        stats: dict[str, dict[str, float]] = {}
        for row in rows:
            stats[str(row.org_id)] = {
                "data_count": float(row.data_count or 0),
                "exceed_count": float(row.exceed_count or 0),
                "invalid_count": float(row.invalid_count or 0),
            }
        return stats

    async def _range_stats_by_org(
        self,
        org_ids: list[str],
        start_date: date,
        end_date: date,
    ) -> dict[str, dict[str, float]]:
        if not org_ids:
            return {}
        query = (
            select(
                MonitoringDailyStats.org_id,
                func.sum(MonitoringDailyStats.data_count).label("data_count"),
                func.sum(MonitoringDailyStats.exceed_count).label("exceed_count"),
                func.sum(MonitoringDailyStats.invalid_count).label("invalid_count"),
            )
            .where(
                and_(
                    MonitoringDailyStats.stat_date >= start_date,
                    MonitoringDailyStats.stat_date <= end_date,
                    MonitoringDailyStats.org_id.in_(org_ids),
                )
            )
            .group_by(MonitoringDailyStats.org_id)
        )
        result = await self.db.execute(query)
        rows = result.all()
        stats: dict[str, dict[str, float]] = {}
        for row in rows:
            stats[str(row.org_id)] = {
                "data_count": float(row.data_count or 0),
                "exceed_count": float(row.exceed_count or 0),
                "invalid_count": float(row.invalid_count or 0),
            }
        return stats

    async def _resolve_industry_by_org(
        self,
        orgs: list[Organization],
        org_ids: list[str],
    ) -> dict[str, str]:
        industry_by_org: dict[str, str] = {}
        for org in orgs:
            if org.industry_type:
                industry_by_org[str(org.id)] = org.industry_type

        missing = [oid for oid in org_ids if oid not in industry_by_org]
        if not missing:
            return industry_by_org

        query = (
            select(
                Device.org_id,
                Device.industry_type,
                func.count(Device.id).label("device_count"),
            )
            .where(
                and_(
                    Device.org_id.in_(missing),
                    Device.industry_type.isnot(None),
                )
            )
            .group_by(Device.org_id, Device.industry_type)
        )
        result = await self.db.execute(query)
        rows = result.all()
        per_org: dict[str, dict[str, int]] = {}
        for row in rows:
            org_id = str(row.org_id)
            per_org.setdefault(org_id, {})
            per_org[org_id][row.industry_type] = int(row.device_count or 0)

        for org_id, counts in per_org.items():
            industry = max(counts.items(), key=lambda kv: kv[1])[0]
            industry_by_org[org_id] = industry

        return industry_by_org

    def _compute_risk(
        self,
        device_stats: dict[str, int],
        daily_stats: dict[str, float],
    ) -> tuple[float, str]:
        total = device_stats.get("total", 0)
        offline_rate = _safe_divide(device_stats.get("offline", 0), total)
        alarm_rate = _safe_divide(device_stats.get("alarm", 0), total)
        exceed_rate = _safe_divide(daily_stats.get("exceed_count", 0), daily_stats.get("data_count", 0))
        invalid_rate = _safe_divide(daily_stats.get("invalid_count", 0), daily_stats.get("data_count", 0))

        score = (
            exceed_rate * 0.4
            + invalid_rate * 0.2
            + offline_rate * 0.2
            + alarm_rate * 0.2
        ) * 100
        level = _risk_level(score)
        return round(score, 2), level

    async def get_overview(
        self,
        user: User,
        target_date: date | None,
        region_code: str | None,
        park_code: str | None,
    ) -> dict[str, Any]:
        target = _resolve_target_date(target_date)
        orgs = await self._fetch_enterprise_orgs(user, region_code, park_code)
        org_ids = [str(org.id) for org in orgs]

        device_stats = await self._device_stats_by_org(org_ids)
        daily_stats = await self._daily_stats_by_org(org_ids, target)
        industry_by_org = await self._resolve_industry_by_org(orgs, org_ids)

        risk_distribution: dict[str, int] = {level: 0 for _, level in RISK_LEVELS}
        industry_counts: dict[str, int] = {}
        region_counts: dict[str, int] = {}

        total_devices = 0
        online_devices = 0
        offline_devices = 0

        for org in orgs:
            org_id = str(org.id)
            d_stats = device_stats.get(org_id, {"total": 0, "online": 0, "offline": 0, "alarm": 0})
            total_devices += d_stats.get("total", 0)
            online_devices += d_stats.get("online", 0)
            offline_devices += d_stats.get("offline", 0)

            score, level = self._compute_risk(d_stats, daily_stats.get(org_id, {}))
            risk_distribution[level] = risk_distribution.get(level, 0) + 1

            industry = industry_by_org.get(org_id, "other")
            industry_counts[industry] = industry_counts.get(industry, 0) + 1

            region_key = org.park_code if park_code or (org.park_code and not org.region_code) else org.region_code
            region_key = region_key or "unknown"
            region_counts[region_key] = region_counts.get(region_key, 0) + 1

        industry_distribution = []
        for industry, count in sorted(industry_counts.items(), key=lambda kv: kv[0]):
            threshold = MIN_SAMPLE_BY_INDUSTRY.get(industry, DEFAULT_MIN_SAMPLE)
            industry_distribution.append({
                "industry": industry,
                "count": count,
                "insufficient": count < threshold,
            })

        region_distribution = [
            {"region_code": key, "count": count}
            for key, count in sorted(region_counts.items(), key=lambda kv: kv[0])
        ]

        return {
            "target_date": target.isoformat(),
            "enterprise_count": len(orgs),
            "device_count": total_devices,
            "online_device_count": online_devices,
            "offline_device_count": offline_devices,
            "risk_distribution": [
                {"level": level, "count": risk_distribution.get(level, 0)}
                for _, level in RISK_LEVELS
            ],
            "industry_distribution": industry_distribution,
            "region_distribution": region_distribution,
        }

    async def get_overview_range(
        self,
        user: User,
        start_date: date,
        end_date: date,
        region_code: str | None,
        park_code: str | None,
    ) -> dict[str, Any]:
        start, end = _resolve_date_range(start_date, end_date)
        orgs = await self._fetch_enterprise_orgs(user, region_code, park_code)
        org_ids = [str(org.id) for org in orgs]

        device_stats = await self._device_stats_by_org(org_ids)
        range_stats = await self._range_stats_by_org(org_ids, start, end)
        industry_by_org = await self._resolve_industry_by_org(orgs, org_ids)

        risk_distribution: dict[str, int] = {level: 0 for _, level in RISK_LEVELS}
        industry_counts: dict[str, int] = {}
        region_counts: dict[str, int] = {}

        total_devices = 0
        online_devices = 0
        offline_devices = 0

        for org in orgs:
            org_id = str(org.id)
            d_stats = device_stats.get(org_id, {"total": 0, "online": 0, "offline": 0, "alarm": 0})
            total_devices += d_stats.get("total", 0)
            online_devices += d_stats.get("online", 0)
            offline_devices += d_stats.get("offline", 0)

            score, level = self._compute_risk(d_stats, range_stats.get(org_id, {}))
            risk_distribution[level] = risk_distribution.get(level, 0) + 1

            industry = industry_by_org.get(org_id, "other")
            industry_counts[industry] = industry_counts.get(industry, 0) + 1

            region_key = org.park_code if park_code or (org.park_code and not org.region_code) else org.region_code
            region_key = region_key or "unknown"
            region_counts[region_key] = region_counts.get(region_key, 0) + 1

        industry_distribution = []
        for industry, count in sorted(industry_counts.items(), key=lambda kv: kv[0]):
            threshold = MIN_SAMPLE_BY_INDUSTRY.get(industry, DEFAULT_MIN_SAMPLE)
            industry_distribution.append({
                "industry": industry,
                "count": count,
                "insufficient": count < threshold,
            })

        region_distribution = [
            {"region_code": key, "count": count}
            for key, count in sorted(region_counts.items(), key=lambda kv: kv[0])
        ]

        return {
            "start_date": start.isoformat(),
            "end_date": end.isoformat(),
            "enterprise_count": len(orgs),
            "device_count": total_devices,
            "online_device_count": online_devices,
            "offline_device_count": offline_devices,
            "risk_distribution": [
                {"level": level, "count": risk_distribution.get(level, 0)}
                for _, level in RISK_LEVELS
            ],
            "industry_distribution": industry_distribution,
            "region_distribution": region_distribution,
        }

    async def get_heatmap(
        self,
        user: User,
        target_date: date | None,
        resolution: int,
        region_code: str | None,
        park_code: str | None,
    ) -> dict[str, Any]:
        target = _resolve_target_date(target_date)
        orgs = await self._fetch_enterprise_orgs(user, region_code, park_code)
        org_ids = [str(org.id) for org in orgs]
        if not org_ids:
            return {"target_date": target.isoformat(), "resolution": resolution, "cells": []}

        device_stats = await self._device_stats_by_org(org_ids)
        daily_stats = await self._daily_stats_by_org(org_ids, target)

        query = select(Device.org_id, Device.latitude, Device.longitude).where(
            and_(
                Device.org_id.in_(org_ids),
                Device.latitude.isnot(None),
                Device.longitude.isnot(None),
            )
        )
        result = await self.db.execute(query)
        rows = result.all()

        cells: dict[str, dict[str, Any]] = {}
        for row in rows:
            lat = float(row.latitude)
            lng = float(row.longitude)
            cell = h3.latlng_to_cell(lat, lng, resolution)
            org_id = str(row.org_id)

            cell_info = cells.setdefault(cell, {
                "org_ids": set(),
                "device_count": 0,
                "risk_scores": [],
            })
            cell_info["device_count"] += 1
            if org_id not in cell_info["org_ids"]:
                d_stats = device_stats.get(org_id, {"total": 0, "online": 0, "offline": 0, "alarm": 0})
                score, _ = self._compute_risk(d_stats, daily_stats.get(org_id, {}))
                cell_info["org_ids"].add(org_id)
                cell_info["risk_scores"].append(score)

        payload_cells = []
        for cell, info in cells.items():
            boundary = h3.cell_to_boundary(cell)
            coords = [[lng, lat] for lat, lng in boundary]
            enterprise_count = len(info["org_ids"])
            avg_score = sum(info["risk_scores"]) / enterprise_count if enterprise_count else 0.0
            payload_cells.append({
                "h3_index": cell,
                "boundary": coords,
                "risk_level": _risk_level(avg_score),
                "risk_score": round(avg_score, 2),
                "enterprise_count": enterprise_count,
                "device_count": info["device_count"],
            })

        return {
            "target_date": target.isoformat(),
            "resolution": resolution,
            "cells": payload_cells,
        }

    async def get_trends(
        self,
        user: User,
        start_date: date | None,
        end_date: date | None,
        granularity: str,
    ) -> dict[str, Any]:
        start, end = _resolve_date_range(start_date, end_date)
        orgs = await self._fetch_enterprise_orgs(user, None, None)
        org_ids = [str(org.id) for org in orgs]
        if not org_ids:
            return {"granularity": granularity, "series": []}

        device_stats = await self._device_stats_by_org(org_ids)

        stats_query = (
            select(
                MonitoringDailyStats.stat_date,
                MonitoringDailyStats.org_id,
                func.sum(MonitoringDailyStats.data_count).label("data_count"),
                func.sum(MonitoringDailyStats.exceed_count).label("exceed_count"),
                func.sum(MonitoringDailyStats.invalid_count).label("invalid_count"),
            )
            .where(
                and_(
                    MonitoringDailyStats.stat_date >= start,
                    MonitoringDailyStats.stat_date <= end,
                    MonitoringDailyStats.org_id.in_(org_ids),
                )
            )
            .group_by(MonitoringDailyStats.stat_date, MonitoringDailyStats.org_id)
        )
        result = await self.db.execute(stats_query)
        rows = result.all()

        stats_by_date: dict[date, dict[str, dict[str, float]]] = {}
        for row in rows:
            stats_by_date.setdefault(row.stat_date, {})[str(row.org_id)] = {
                "data_count": float(row.data_count or 0),
                "exceed_count": float(row.exceed_count or 0),
                "invalid_count": float(row.invalid_count or 0),
            }

        series = []
        cursor = start
        while cursor <= end:
            distribution: dict[str, int] = {level: 0 for _, level in RISK_LEVELS}
            per_org = stats_by_date.get(cursor, {})
            for org_id in org_ids:
                d_stats = device_stats.get(org_id, {"total": 0, "online": 0, "offline": 0, "alarm": 0})
                score, level = self._compute_risk(d_stats, per_org.get(org_id, {}))
                distribution[level] = distribution.get(level, 0) + 1
            series.append({
                "date": cursor.isoformat(),
                "risk_distribution": [
                    {"level": level, "count": distribution.get(level, 0)}
                    for _, level in RISK_LEVELS
                ],
            })
            cursor += timedelta(days=1)

        if granularity == "monthly":
            monthly: dict[str, dict[str, int]] = {}
            for item in series:
                month_key = item["date"][:7]
                bucket = monthly.setdefault(month_key, {level: 0 for _, level in RISK_LEVELS})
                for entry in item["risk_distribution"]:
                    bucket[entry["level"]] += entry["count"]
            series = [
                {
                    "date": month,
                    "risk_distribution": [
                        {"level": level, "count": counts.get(level, 0)}
                        for _, level in RISK_LEVELS
                    ],
                }
                for month, counts in sorted(monthly.items())
            ]

        return {"granularity": granularity, "series": series}

    async def get_consistency(
        self,
        user: User,
        start_date: date | None,
        end_date: date | None,
    ) -> dict[str, Any]:
        start, end = _resolve_date_range(start_date, end_date, default_days=90)
        orgs = await self._fetch_enterprise_orgs(user, None, None)
        org_ids = [str(org.id) for org in orgs]
        if not org_ids:
            return {"summary": {"high": 0, "medium": 0, "low": 0}, "industry_breakdown": [], "region_breakdown": []}

        industry_by_org = await self._resolve_industry_by_org(orgs, org_ids)
        region_by_org = {str(org.id): (org.region_code or "unknown") for org in orgs}

        reports_query = (
            select(
                SelfInspectionReport.org_id,
                SelfInspectionReport.inspection_date,
                SelfInspectionData.pollutant_code,
                SelfInspectionData.value,
            )
            .join(SelfInspectionData, SelfInspectionReport.id == SelfInspectionData.report_id)
            .where(
                and_(
                    SelfInspectionReport.inspection_date >= start,
                    SelfInspectionReport.inspection_date <= end,
                    SelfInspectionReport.org_id.in_([org.id for org in orgs]),
                )
            )
        )
        report_rows = (await self.db.execute(reports_query)).all()

        stats_query = (
            select(
                MonitoringDailyStats.org_id,
                MonitoringDailyStats.stat_date,
                MonitoringDailyStats.pollutant_code,
                MonitoringDailyStats.avg_value,
            )
            .where(
                and_(
                    MonitoringDailyStats.stat_date >= start,
                    MonitoringDailyStats.stat_date <= end,
                    MonitoringDailyStats.org_id.in_(org_ids),
                )
            )
        )
        stats_rows = (await self.db.execute(stats_query)).all()

        avg_lookup: dict[tuple[str, date, str], float] = {}
        for row in stats_rows:
            if row.avg_value is None:
                continue
            avg_lookup[(str(row.org_id), row.stat_date, row.pollutant_code)] = float(row.avg_value)

        per_org: dict[str, dict[str, int]] = {}
        for row in report_rows:
            org_id = str(row.org_id)
            value = _parse_numeric_value(row.value)
            if value is None:
                continue
            key = (org_id, row.inspection_date, row.pollutant_code)
            if key not in avg_lookup:
                continue
            avg_value = avg_lookup[key]
            if avg_value <= 0:
                continue
            diff_ratio = abs(value - avg_value) / avg_value
            consistent = diff_ratio <= 0.3
            stats = per_org.setdefault(org_id, {"total": 0, "consistent": 0})
            stats["total"] += 1
            if consistent:
                stats["consistent"] += 1

        summary = {"high": 0, "medium": 0, "low": 0}
        industry_breakdown: dict[str, dict[str, int]] = {}
        region_breakdown: dict[str, dict[str, int]] = {}

        for org_id, stats in per_org.items():
            total = stats["total"]
            if total <= 0:
                continue
            score = stats["consistent"] / total
            if score >= 0.8:
                bucket = "high"
            elif score >= 0.5:
                bucket = "medium"
            else:
                bucket = "low"
            summary[bucket] += 1

            industry = industry_by_org.get(org_id, "other")
            industry_breakdown.setdefault(industry, {"high": 0, "medium": 0, "low": 0})
            industry_breakdown[industry][bucket] += 1

            region = region_by_org.get(org_id, "unknown")
            region_breakdown.setdefault(region, {"high": 0, "medium": 0, "low": 0})
            region_breakdown[region][bucket] += 1

        industry_payload = [
            {"industry": key, **counts}
            for key, counts in sorted(industry_breakdown.items())
        ]
        region_payload = [
            {"region_code": key, **counts}
            for key, counts in sorted(region_breakdown.items())
        ]

        return {
            "summary": summary,
            "industry_breakdown": industry_payload,
            "region_breakdown": region_payload,
        }

    async def build_report_payload(
        self,
        user: User,
        report_type: str,
        target_date: date | None,
        year: int | None,
        month: int | None,
        region_code: str | None,
        park_code: str | None,
    ) -> dict[str, Any]:
        report_type = report_type.lower()
        if report_type not in {"daily", "monthly"}:
            raise ValueError("invalid report_type")

        if report_type == "daily":
            target = _resolve_target_date(target_date)
            overview = await self.get_overview(user, target, region_code, park_code)
            period_label = target.isoformat()
            return {
                "report_type": "daily",
                "period_label": period_label,
                "start_date": target.isoformat(),
                "end_date": target.isoformat(),
                "overview": overview,
            }

        if year is None or month is None:
            now = datetime.now(ZoneInfo("Asia/Shanghai")).date()
            year = now.year
            month = now.month

        start, end = _resolve_month_range(year, month)
        end = _resolve_target_date(end)
        if start > end:
            start = end
        overview = await self.get_overview_range(user, start, end, region_code, park_code)
        period_label = f"{start.strftime('%Y-%m')}"
        return {
            "report_type": "monthly",
            "period_label": period_label,
            "start_date": start.isoformat(),
            "end_date": end.isoformat(),
            "overview": overview,
        }

    async def generate_excel_report(self, payload: dict[str, Any]) -> bytes:
        overview = payload["overview"]

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        header_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_text = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        ws.merge_cells("A1:D1")
        ws["A1"] = "Regulator Summary Report"
        ws["A1"].font = header_font
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A3"] = "Report Type"
        ws["B3"] = payload["report_type"]
        ws["A4"] = "Period"
        ws["B4"] = payload["period_label"]
        ws["A5"] = "Start Date"
        ws["B5"] = payload["start_date"]
        ws["A6"] = "End Date"
        ws["B6"] = payload["end_date"]

        summary_rows = [
            ("Enterprise Count", overview["enterprise_count"]),
            ("Device Count", overview["device_count"]),
            ("Online Devices", overview["online_device_count"]),
            ("Offline Devices", overview["offline_device_count"]),
        ]

        ws["A8"] = "Summary"
        ws["A8"].font = section_font
        row = 9
        for label, value in summary_rows:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1

        row += 1
        ws[f"A{row}"] = "Risk Distribution"
        ws[f"A{row}"].font = section_font
        row += 1
        ws[f"A{row}"] = "Level"
        ws[f"B{row}"] = "Count"
        ws[f"A{row}"].font = header_text
        ws[f"B{row}"].font = header_text
        ws[f"A{row}"].fill = header_fill
        ws[f"B{row}"].fill = header_fill
        ws[f"A{row}"].border = thin_border
        ws[f"B{row}"].border = thin_border
        row += 1
        for entry in overview["risk_distribution"]:
            ws[f"A{row}"] = entry["level"]
            ws[f"B{row}"] = entry["count"]
            ws[f"A{row}"].border = thin_border
            ws[f"B{row}"].border = thin_border
            row += 1

        row += 1
        ws[f"A{row}"] = "Industry Distribution"
        ws[f"A{row}"].font = section_font
        row += 1
        ws[f"A{row}"] = "Industry"
        ws[f"B{row}"] = "Count"
        ws[f"C{row}"] = "Insufficient"
        for col in ("A", "B", "C"):
            ws[f"{col}{row}"].font = header_text
            ws[f"{col}{row}"].fill = header_fill
            ws[f"{col}{row}"].border = thin_border
        row += 1
        for entry in overview["industry_distribution"]:
            ws[f"A{row}"] = entry["industry"]
            ws[f"B{row}"] = entry["count"]
            ws[f"C{row}"] = "yes" if entry["insufficient"] else "no"
            for col in ("A", "B", "C"):
                ws[f"{col}{row}"].border = thin_border
            row += 1

        row += 1
        ws[f"A{row}"] = "Region Distribution"
        ws[f"A{row}"].font = section_font
        row += 1
        ws[f"A{row}"] = "Region"
        ws[f"B{row}"] = "Count"
        for col in ("A", "B"):
            ws[f"{col}{row}"].font = header_text
            ws[f"{col}{row}"].fill = header_fill
            ws[f"{col}{row}"].border = thin_border
        row += 1
        for entry in overview["region_distribution"]:
            ws[f"A{row}"] = entry["region_code"]
            ws[f"B{row}"] = entry["count"]
            ws[f"A{row}"].border = thin_border
            ws[f"B{row}"].border = thin_border
            row += 1

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    async def generate_pdf_report(self, payload: dict[str, Any]) -> bytes:
        overview = payload["overview"]

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=18 * mm,
            leftMargin=18 * mm,
            topMargin=16 * mm,
            bottomMargin=16 * mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "RegulatorTitle",
            parent=styles["Heading1"],
            fontSize=16,
            alignment=1,
            spaceAfter=12,
        )
        section_style = ParagraphStyle(
            "RegulatorSection",
            parent=styles["Heading2"],
            fontSize=12,
            spaceAfter=6,
        )
        normal_style = ParagraphStyle(
            "RegulatorNormal",
            parent=styles["Normal"],
            fontSize=9,
            spaceAfter=4,
        )

        elements: list[Any] = []
        elements.append(Paragraph("Regulator Summary Report", title_style))
        elements.append(Paragraph(f"Report Type: {payload['report_type']}", normal_style))
        elements.append(Paragraph(f"Period: {payload['period_label']}", normal_style))
        elements.append(Paragraph(f"Start Date: {payload['start_date']}", normal_style))
        elements.append(Paragraph(f"End Date: {payload['end_date']}", normal_style))
        elements.append(Spacer(1, 8))

        summary_data = [
            ["Metric", "Value"],
            ["Enterprise Count", str(overview["enterprise_count"])],
            ["Device Count", str(overview["device_count"])],
            ["Online Devices", str(overview["online_device_count"])],
            ["Offline Devices", str(overview["offline_device_count"])],
        ]
        summary_table = Table(summary_data, colWidths=[120, 120])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(Paragraph("Summary", section_style))
        elements.append(summary_table)
        elements.append(Spacer(1, 10))

        risk_data = [["Level", "Count"]]
        for entry in overview["risk_distribution"]:
            risk_data.append([entry["level"], str(entry["count"])])
        risk_table = Table(risk_data, colWidths=[80, 80])
        risk_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(Paragraph("Risk Distribution", section_style))
        elements.append(risk_table)
        elements.append(Spacer(1, 10))

        industry_data = [["Industry", "Count", "Insufficient"]]
        for entry in overview["industry_distribution"]:
            industry_data.append([
                entry["industry"],
                str(entry["count"]),
                "yes" if entry["insufficient"] else "no",
            ])
        industry_table = Table(industry_data, colWidths=[160, 60, 80])
        industry_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(Paragraph("Industry Distribution", section_style))
        elements.append(industry_table)
        elements.append(Spacer(1, 10))

        region_data = [["Region", "Count"]]
        for entry in overview["region_distribution"]:
            region_data.append([entry["region_code"], str(entry["count"])])
        region_table = Table(region_data, colWidths=[160, 60])
        region_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(Paragraph("Region Distribution", section_style))
        elements.append(region_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
