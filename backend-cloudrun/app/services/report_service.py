from __future__ import annotations

"""Report generation service for daily and monthly reports."""

import io
from datetime import datetime, date, timedelta
from typing import Any
from uuid import UUID

import pandas as pd
import structlog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.monitoring_service import MonitoringService
from app.protocols.enums import PARAMETER_DESCRIPTIONS, PARAMETER_UNITS

logger = structlog.get_logger()


# Register Chinese CID font for PDF (built-in support, no external font file needed)
try:
    pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
    CHINESE_FONT = 'STSong-Light'
    logger.info("Registered Chinese CID font: STSong-Light")
except Exception as e:
    CHINESE_FONT = 'Helvetica'
    logger.warning(f"Chinese font registration failed: {e}, using Helvetica as fallback")


class ReportService:
    """Service for generating environmental monitoring reports."""

    def __init__(self, db_session: AsyncSession):
        """
        初始化报告服务。

        Args:
            db_session: 数据库会话，用于查询监测数据
        """
        self.db_session = db_session
        self.monitoring_service = MonitoringService(db_session)

    async def get_report_statistics(
        self,
        device_id: str,
        device_name: str,
        start_time: datetime,
        end_time: datetime,
        pollutant_codes: list[str] | None = None,
        thresholds: dict[str, float] | None = None,
    ) -> dict[str, Any]:
        """
        Generate statistics for report preview.

        Args:
            device_id: Device MN number
            device_name: Device display name
            start_time: Start of the reporting period
            end_time: End of the reporting period
            pollutant_codes: List of pollutant codes to include (optional)
            thresholds: Dict of pollutant_code -> alarm_value for exceedance calculation

        Returns:
            Dict containing statistics for each pollutant
        """
        # Query all data for the period (使用 MySQL MonitoringService)
        data = await self.monitoring_service.query_monitoring_data(
            device_id=device_id,
            start_time=start_time,
            end_time=end_time,
            limit=100000,  # Large limit for full data
        )

        if not data:
            return {
                "device_id": device_id,
                "device_name": device_name,
                "period": {
                    "start": start_time.isoformat(),
                    "end": end_time.isoformat(),
                    "days": (end_time - start_time).days,
                },
                "pollutants": [],
                "summary": {
                    "total_records": 0,
                    "expected_records": 0,
                    "capture_rate": 0,
                    "exceedance_count": 0,
                },
            }

        # Group data by pollutant
        df = pd.DataFrame(data)
        pollutant_stats = []

        # Filter pollutants if specified
        if pollutant_codes:
            unique_pollutants = [p for p in df["pollutant_code"].unique() if p in pollutant_codes]
        else:
            unique_pollutants = df["pollutant_code"].unique()

        total_exceedance = 0

        for pollutant_code in unique_pollutants:
            pollutant_data = df[df["pollutant_code"] == pollutant_code]
            values = pollutant_data["value"].astype(float)

            # Calculate statistics
            min_val = float(values.min())
            max_val = float(values.max())
            avg_val = float(values.mean())
            std_val = float(values.std()) if len(values) > 1 else 0.0
            count = len(values)

            # Calculate exceedance count
            threshold = thresholds.get(pollutant_code) if thresholds else None
            exceedance_count = 0
            if threshold:
                exceedance_count = int((values > threshold).sum())
                total_exceedance += exceedance_count

            # Count abnormal flags
            abnormal_flags = pollutant_data["flag"].isin(["A", "F", "D", "S"]).sum()

            pollutant_stats.append({
                "pollutant_code": pollutant_code,
                "pollutant_name": PARAMETER_DESCRIPTIONS.get(pollutant_code, pollutant_code),
                "unit": PARAMETER_UNITS.get(pollutant_code, ""),
                "min_value": round(min_val, 4),
                "max_value": round(max_val, 4),
                "avg_value": round(avg_val, 4),
                "std_value": round(std_val, 4),
                "data_count": count,
                "exceedance_count": exceedance_count,
                "threshold": threshold,
                "abnormal_flag_count": int(abnormal_flags),
            })

        # Calculate overall capture rate (assuming minute data, 1440 records/day per pollutant)
        period_days = max((end_time - start_time).days, 1)
        expected_records = period_days * 1440 * len(unique_pollutants)
        actual_records = len(df)
        capture_rate = min(actual_records / expected_records * 100, 100) if expected_records > 0 else 0

        return {
            "device_id": device_id,
            "device_name": device_name,
            "period": {
                "start": start_time.isoformat(),
                "end": end_time.isoformat(),
                "days": period_days,
            },
            "pollutants": pollutant_stats,
            "summary": {
                "total_records": actual_records,
                "expected_records": expected_records,
                "capture_rate": round(capture_rate, 2),
                "exceedance_count": total_exceedance,
            },
        }

    async def generate_daily_report(
        self,
        device_id: str,
        device_name: str,
        report_date: date,
        pollutant_codes: list[str] | None = None,
        thresholds: dict[str, float] | None = None,
    ) -> dict[str, Any]:
        """
        Generate daily report data.

        Args:
            device_id: Device MN number
            device_name: Device display name
            report_date: The date to generate report for
            pollutant_codes: List of pollutant codes to include
            thresholds: Dict of pollutant_code -> alarm_value

        Returns:
            Dict containing daily report data
        """
        start_time = datetime.combine(report_date, datetime.min.time())
        end_time = datetime.combine(report_date, datetime.max.time())

        return await self.get_report_statistics(
            device_id=device_id,
            device_name=device_name,
            start_time=start_time,
            end_time=end_time,
            pollutant_codes=pollutant_codes,
            thresholds=thresholds,
        )

    async def generate_monthly_report(
        self,
        device_id: str,
        device_name: str,
        year: int,
        month: int,
        pollutant_codes: list[str] | None = None,
        thresholds: dict[str, float] | None = None,
    ) -> dict[str, Any]:
        """
        Generate monthly report data.

        Args:
            device_id: Device MN number
            device_name: Device display name
            year: Report year
            month: Report month (1-12)
            pollutant_codes: List of pollutant codes to include
            thresholds: Dict of pollutant_code -> alarm_value

        Returns:
            Dict containing monthly report data
        """
        start_time = datetime(year, month, 1)
        if month == 12:
            end_time = datetime(year + 1, 1, 1) - timedelta(seconds=1)
        else:
            end_time = datetime(year, month + 1, 1) - timedelta(seconds=1)

        return await self.get_report_statistics(
            device_id=device_id,
            device_name=device_name,
            start_time=start_time,
            end_time=end_time,
            pollutant_codes=pollutant_codes,
            thresholds=thresholds,
        )

    async def generate_excel_report(
        self,
        device_id: str,
        device_name: str,
        start_time: datetime,
        end_time: datetime,
        report_type: str = "daily",
        pollutant_codes: list[str] | None = None,
        thresholds: dict[str, float] | None = None,
    ) -> bytes:
        """
        Generate Excel report file.

        Args:
            device_id: Device MN number
            device_name: Device display name
            start_time: Start of the reporting period
            end_time: End of the reporting period
            report_type: 'daily' or 'monthly'
            pollutant_codes: List of pollutant codes to include
            thresholds: Dict of pollutant_code -> alarm_value

        Returns:
            Excel file as bytes
        """
        # Get statistics
        stats = await self.get_report_statistics(
            device_id=device_id,
            device_name=device_name,
            start_time=start_time,
            end_time=end_time,
            pollutant_codes=pollutant_codes,
            thresholds=thresholds,
        )

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "统计摘要"

        # Styles
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_text = Font(bold=True, color="FFFFFF")

        # Title
        report_title = "日报" if report_type == "daily" else "月报"
        ws.merge_cells('A1:H1')
        ws['A1'] = f"环境监测数据{report_title}"
        ws['A1'].font = header_font
        ws['A1'].alignment = Alignment(horizontal='center')

        # Device info
        ws['A3'] = "设备名称:"
        ws['B3'] = device_name
        ws['A4'] = "设备MN号:"
        ws['B4'] = device_id
        ws['A5'] = "报告周期:"
        ws['B5'] = f"{start_time.strftime('%Y-%m-%d %H:%M')} 至 {end_time.strftime('%Y-%m-%d %H:%M')}"
        ws['A6'] = "生成时间:"
        ws['B6'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Summary statistics
        ws['A8'] = "数据概要"
        ws['A8'].font = subheader_font
        ws['A9'] = "总记录数:"
        ws['B9'] = stats['summary']['total_records']
        ws['A10'] = "数据捕获率:"
        ws['B10'] = f"{stats['summary']['capture_rate']}%"
        ws['A11'] = "超标次数:"
        ws['B11'] = stats['summary']['exceedance_count']

        # Pollutant statistics table
        ws['A13'] = "各污染物统计"
        ws['A13'].font = subheader_font

        # Table headers
        headers = ["污染物代码", "污染物名称", "单位", "最小值", "最大值", "平均值", "标准差", "数据量", "超标次数"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=14, column=col, value=header)
            cell.font = header_text
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Table data
        for row_idx, pollutant in enumerate(stats['pollutants'], 15):
            data = [
                pollutant['pollutant_code'],
                pollutant['pollutant_name'],
                pollutant['unit'],
                pollutant['min_value'],
                pollutant['max_value'],
                pollutant['avg_value'],
                pollutant['std_value'],
                pollutant['data_count'],
                pollutant['exceedance_count'],
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10

        # Add detail data sheet
        ws_detail = wb.create_sheet(title="详细数据")

        # Get raw data (使用 MySQL MonitoringService)
        raw_data = await self.monitoring_service.query_monitoring_data(
            device_id=device_id,
            start_time=start_time,
            end_time=end_time,
            limit=100000,
        )

        if raw_data:
            # Convert to DataFrame for better handling
            df = pd.DataFrame(raw_data)

            # Add pollutant name column
            df['pollutant_name'] = df['pollutant_code'].map(
                lambda x: PARAMETER_DESCRIPTIONS.get(x, x)
            )

            # Sort by timestamp
            df = df.sort_values('ts')

            # Headers for detail sheet
            detail_headers = ["时间", "污染物代码", "污染物名称", "数值", "标志", "状态"]
            for col, header in enumerate(detail_headers, 1):
                cell = ws_detail.cell(row=1, column=col, value=header)
                cell.font = header_text
                cell.fill = header_fill
                cell.border = thin_border

            # Data
            for row_idx, (_, row) in enumerate(df.iterrows(), 2):
                ts = row['ts']
                if isinstance(ts, str):
                    ts_str = ts
                else:
                    ts_str = ts.strftime('%Y-%m-%d %H:%M:%S') if hasattr(ts, 'strftime') else str(ts)

                data = [
                    ts_str,
                    row['pollutant_code'],
                    row['pollutant_name'],
                    row['value'],
                    row['flag'],
                    row['status'],
                ]
                for col, value in enumerate(data, 1):
                    cell = ws_detail.cell(row=row_idx, column=col, value=value)
                    cell.border = thin_border

            # Adjust column widths
            ws_detail.column_dimensions['A'].width = 20
            ws_detail.column_dimensions['B'].width = 15
            ws_detail.column_dimensions['C'].width = 20
            ws_detail.column_dimensions['D'].width = 12
            ws_detail.column_dimensions['E'].width = 8
            ws_detail.column_dimensions['F'].width = 8

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    async def generate_pdf_report(
        self,
        device_id: str,
        device_name: str,
        start_time: datetime,
        end_time: datetime,
        report_type: str = "daily",
        pollutant_codes: list[str] | None = None,
        thresholds: dict[str, float] | None = None,
    ) -> bytes:
        """
        Generate PDF report file.

        Args:
            device_id: Device MN number
            device_name: Device display name
            start_time: Start of the reporting period
            end_time: End of the reporting period
            report_type: 'daily' or 'monthly'
            pollutant_codes: List of pollutant codes to include
            thresholds: Dict of pollutant_code -> alarm_value

        Returns:
            PDF file as bytes
        """
        # Get statistics
        stats = await self.get_report_statistics(
            device_id=device_id,
            device_name=device_name,
            start_time=start_time,
            end_time=end_time,
            pollutant_codes=pollutant_codes,
            thresholds=thresholds,
        )

        # Create PDF buffer
        buffer = io.BytesIO()

        # Create PDF document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=20*mm,
            leftMargin=20*mm,
            topMargin=20*mm,
            bottomMargin=20*mm,
        )

        # Build content
        elements = []

        # Styles
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=CHINESE_FONT,
            fontSize=18,
            alignment=1,  # Center
            spaceAfter=20,
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontName=CHINESE_FONT,
            fontSize=14,
            spaceAfter=10,
        )

        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontName=CHINESE_FONT,
            fontSize=10,
            spaceAfter=5,
        )

        # Title
        report_title = "日报" if report_type == "daily" else "月报"
        elements.append(Paragraph(f"环境监测数据{report_title}", title_style))
        elements.append(Spacer(1, 10))

        # Device info
        elements.append(Paragraph(f"设备名称: {device_name}", normal_style))
        elements.append(Paragraph(f"设备MN号: {device_id}", normal_style))
        elements.append(Paragraph(
            f"报告周期: {start_time.strftime('%Y-%m-%d %H:%M')} 至 {end_time.strftime('%Y-%m-%d %H:%M')}",
            normal_style
        ))
        elements.append(Paragraph(
            f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            normal_style
        ))
        elements.append(Spacer(1, 20))

        # Summary
        elements.append(Paragraph("数据概要", heading_style))
        summary_data = [
            ["项目", "数值"],
            ["总记录数", str(stats['summary']['total_records'])],
            ["数据捕获率", f"{stats['summary']['capture_rate']}%"],
            ["超标次数", str(stats['summary']['exceedance_count'])],
        ]

        summary_table = Table(summary_data, colWidths=[100, 100])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), CHINESE_FONT),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Pollutant statistics
        elements.append(Paragraph("各污染物统计", heading_style))

        if stats['pollutants']:
            # Table header
            pollutant_data = [
                ["污染物", "名称", "单位", "最小值", "最大值", "平均值", "数据量", "超标"]
            ]

            # Table data
            for p in stats['pollutants']:
                pollutant_data.append([
                    p['pollutant_code'],
                    p['pollutant_name'][:6],  # Truncate long names
                    p['unit'],
                    f"{p['min_value']:.2f}",
                    f"{p['max_value']:.2f}",
                    f"{p['avg_value']:.2f}",
                    str(p['data_count']),
                    str(p['exceedance_count']),
                ])

            col_widths = [60, 60, 40, 50, 50, 50, 45, 35]
            pollutant_table = Table(pollutant_data, colWidths=col_widths)
            pollutant_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), CHINESE_FONT),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#E8F0FE')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#E8F0FE'), colors.white]),
            ]))
            elements.append(pollutant_table)
        else:
            elements.append(Paragraph("暂无数据", normal_style))

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()


def get_report_service(db_session: AsyncSession) -> ReportService:
    """
    获取 ReportService 实例。

    Args:
        db_session: 数据库会话

    Returns:
        ReportService 实例
    """
    return ReportService(db_session)
