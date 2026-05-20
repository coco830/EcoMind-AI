from __future__ import annotations

"""Regulator aggregation service (T+1, aggregate-only)."""

import io
import json
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Any
from zoneinfo import ZoneInfo

import h3
import structlog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from sqlalchemy import select, func, and_, case
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.device import Device, DeviceStatus
from app.models.monitoring_mysql import MonitoringDailyStats
from app.models.organization import Organization, OrganizationType, OrganizationStatus
from app.models.invitation import InvitationCode
from app.models.regulator_brief import RegulatorBriefUsage
from app.models.self_inspection import SelfInspectionReport, SelfInspectionData
from app.models.user import User

logger = structlog.get_logger(__name__)

RISK_LEVELS = [
    (80, "L5"),
    (60, "L4"),
    (40, "L3"),
    (20, "L2"),
    (0, "L1"),
]

BRIEF_TOP_N = 5
BRIEF_LIMITS = {"daily": 3, "monthly": 1}

DEFAULT_MIN_SAMPLE = 10
MIN_SAMPLE_BY_INDUSTRY: dict[str, int] = {
    "steel": 8,
    "petrochemical": 6,
    "chemical": 6,
    "cement": 6,
    "coking": 6,
    "nonferrous_metal": 6,
    "glass_ceramic": 6,
    "paper_making": 6,
    "textile_dyeing": 6,
    "electroplating": 6,
}


# Register Chinese CID font for PDF output
try:
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    CHINESE_FONT = "STSong-Light"
    logger.info("Registered Chinese CID font: STSong-Light")
except Exception as exc:
    CHINESE_FONT = "Helvetica"
    logger.warning("Chinese font registration failed, fallback to Helvetica", error=str(exc))


class RateLimitError(Exception):
    """Raised when brief generation exceeds rate limits."""

    def __init__(self, period: str, limit: int) -> None:
        self.period = period
        self.limit = limit
        super().__init__(f"limit exceeded for {period}")


@dataclass
class RegulatorScope:
    level: str | None
    codes: list[str]


def _resolve_target_date(target_date: date | None) -> date:
    now = datetime.now(ZoneInfo("Asia/Shanghai")).date()
    latest = now - timedelta(days=1)
    if target_date is None or target_date > latest:
        return latest
    return target_date


def _resolve_date_range(
    start_date: date | None,
    end_date: date | None,
    default_days: int = 30,
) -> tuple[date, date]:
    end = _resolve_target_date(end_date)
    start = start_date or (end - timedelta(days=default_days - 1))
    if start > end:
        start = end
    return start, end


def _resolve_month_range(year: int, month: int) -> tuple[date, date]:
    start = date(year, month, 1)
    if month == 12:
        end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(year, month + 1, 1) - timedelta(days=1)
    return start, end


def _safe_divide(numerator: float, denominator: float) -> float:
    if denominator <= 0:
        return 0.0
    return numerator / denominator


def _risk_level(score: float) -> str:
    for threshold, level in RISK_LEVELS:
        if score >= threshold:
            return level
    return "L1"


def _parse_json_list(value: str | None) -> list[str]:
    if not value:
        return []
    try:
        data = json.loads(value)
    except json.JSONDecodeError:
        return []
    if not isinstance(data, list):
        return []
    return [str(item) for item in data]


def _parse_numeric_value(raw: str) -> float | None:
    if not raw:
        return None
    cleaned = raw.strip().replace(" ", "")
    cleaned = cleaned.replace("×", "x").replace("X", "x")
    if "x10^" in cleaned:
        base, exp = cleaned.split("x10^", 1)
        try:
            return float(base) * (10 ** int(exp))
        except ValueError:
            return None
    if "x10" in cleaned:
        base, exp = cleaned.split("x10", 1)
        try:
            return float(base) * (10 ** int(exp))
        except ValueError:
            return None
    try:
        return float(cleaned)
    except ValueError:
        return None


class RegulatoryService:
    """Aggregate regulator data with strict masking."""

    def __init__(self, db: AsyncSession):
        self.db = db

    def _get_scope(self, user: User) -> RegulatorScope | None:
        if user.is_superadmin:
            return None
        org = getattr(user, "organization", None)
        if not org or getattr(org, "org_type", None) != OrganizationType.REGULATOR.value:
            return RegulatorScope(level=None, codes=[])
        return RegulatorScope(
            level=org.jurisdiction_level,
            codes=_parse_json_list(org.jurisdiction_codes),
        )

    def _apply_org_scope(
        self,
        query,
        scope: RegulatorScope | None,
        region_code: str | None,
        park_code: str | None,
    ):
        if scope and scope.codes:
            if scope.level == "park":
                query = query.where(Organization.park_code.in_(scope.codes))
            else:
                query = query.where(Organization.region_code.in_(scope.codes))
        if region_code:
            query = query.where(Organization.region_code == region_code)
        if park_code:
            query = query.where(Organization.park_code == park_code)
        return query

    async def _fetch_enterprise_orgs(
        self,
        user: User,
        region_code: str | None,
        park_code: str | None,
    ) -> list[Organization]:
        scope = self._get_scope(user)
        if scope and not scope.codes:
            return []
        query = (
            select(Organization)
            .join(InvitationCode, InvitationCode.org_id == Organization.id)
            .where(
                Organization.org_type == OrganizationType.ENTERPRISE.value,
                Organization.status == OrganizationStatus.ACTIVE.value,
            )
            .distinct()
        )
        query = self._apply_org_scope(query, scope, region_code, park_code)
        result = await self.db.execute(query)
        return result.scalars().all()

    async def _device_stats_by_org(self, org_ids: list[str]) -> dict[str, dict[str, int]]:
        if not org_ids:
            return {}
        query = (
            select(
                Device.org_id,
                func.count(Device.id).label("total"),
                func.sum(
                    case((Device.status == DeviceStatus.ONLINE.value, 1), else_=0)
                ).label("online"),
                func.sum(
                    case((Device.status == DeviceStatus.OFFLINE.value, 1), else_=0)
                ).label("offline"),
                func.sum(
                    case((Device.status == DeviceStatus.ALARM.value, 1), else_=0)
                ).label("alarm"),
            )
            .where(Device.org_id.in_(org_ids))
            .group_by(Device.org_id)
        )
        result = await self.db.execute(query)
        rows = result.all()
        stats: dict[str, dict[str, int]] = {}
        for row in rows:
            stats[str(row.org_id)] = {
                "total": int(row.total or 0),
                "online": int(row.online or 0),
                "offline": int(row.offline or 0),
                "alarm": int(row.alarm or 0),
            }
        return stats

    async def _daily_stats_by_org(
        self,
        org_ids: list[str],
        target_date: date,
    ) -> dict[str, dict[str, float]]:
        if not org_ids:
            return {}
        query = (
            select(
                MonitoringDailyStats.org_id,
                func.sum(MonitoringDailyStats.data_count).label("data_count"),
                func.sum(MonitoringDailyStats.exceed_count).label("exceed_count"),
                func.sum(MonitoringDailyStats.invalid_count).label("invalid_count"),
            )
            .where(
                and_(
                    MonitoringDailyStats.stat_date == target_date,
                    MonitoringDailyStats.org_id.in_(org_ids),
                )
            )
            .group_by(MonitoringDailyStats.org_id)
        )
        result = await self.db.execute(query)
        rows = result.all()
        stats: dict[str, dict[str, float]] = {}
        for row in rows:
            stats[str(row.org_id)] = {
                "data_count": float(row.data_count or 0),
                "exceed_count": float(row.exceed_count or 0),
                "invalid_count": float(row.invalid_count or 0),
            }
        return stats

    async def _range_stats_by_org(
        self,
        org_ids: list[str],
        start_date: date,
        end_date: date,
    ) -> dict[str, dict[str, float]]:
        if not org_ids:
            return {}
        query = (
            select(
                MonitoringDailyStats.org_id,
                func.sum(MonitoringDailyStats.data_count).label("data_count"),
                func.sum(MonitoringDailyStats.exceed_count).label("exceed_count"),
                func.sum(MonitoringDailyStats.invalid_count).label("invalid_count"),
            )
            .where(
                and_(
                    MonitoringDailyStats.stat_date >= start_date,
                    MonitoringDailyStats.stat_date <= end_date,
                    MonitoringDailyStats.org_id.in_(org_ids),
                )
            )
            .group_by(MonitoringDailyStats.org_id)
        )
        result = await self.db.execute(query)
        rows = result.all()
        stats: dict[str, dict[str, float]] = {}
        for row in rows:
            stats[str(row.org_id)] = {
                "data_count": float(row.data_count or 0),
                "exceed_count": float(row.exceed_count or 0),
                "invalid_count": float(row.invalid_count or 0),
            }
        return stats

    async def _resolve_industry_by_org(
        self,
        orgs: list[Organization],
        org_ids: list[str],
    ) -> dict[str, str]:
        industry_by_org: dict[str, str] = {}
        for org in orgs:
            if org.industry_type:
                industry_by_org[str(org.id)] = org.industry_type

        missing = [oid for oid in org_ids if oid not in industry_by_org]
        if not missing:
            return industry_by_org

        query = (
            select(
                Device.org_id,
                Device.industry_type,
                func.count(Device.id).label("device_count"),
            )
            .where(
                and_(
                    Device.org_id.in_(missing),
                    Device.industry_type.isnot(None),
                )
            )
            .group_by(Device.org_id, Device.industry_type)
        )
        result = await self.db.execute(query)
        rows = result.all()
        per_org: dict[str, dict[str, int]] = {}
        for row in rows:
            org_id = str(row.org_id)
            per_org.setdefault(org_id, {})
            per_org[org_id][row.industry_type] = int(row.device_count or 0)

        for org_id, counts in per_org.items():
            industry = max(counts.items(), key=lambda kv: kv[1])[0]
            industry_by_org[org_id] = industry

        return industry_by_org

    def _compute_risk(
        self,
        device_stats: dict[str, int],
        daily_stats: dict[str, float],
    ) -> tuple[float, str]:
        total = device_stats.get("total", 0)
        offline_rate = _safe_divide(device_stats.get("offline", 0), total)
        alarm_rate = _safe_divide(device_stats.get("alarm", 0), total)
        exceed_rate = _safe_divide(daily_stats.get("exceed_count", 0), daily_stats.get("data_count", 0))
        invalid_rate = _safe_divide(daily_stats.get("invalid_count", 0), daily_stats.get("data_count", 0))

        score = (
            exceed_rate * 0.4
            + invalid_rate * 0.2
            + offline_rate * 0.2
            + alarm_rate * 0.2
        ) * 100
        level = _risk_level(score)
        return round(score, 2), level

    async def get_overview(
        self,
        user: User,
        target_date: date | None,
        region_code: str | None,
        park_code: str | None,
    ) -> dict[str, Any]:
        target = _resolve_target_date(target_date)
        orgs = await self._fetch_enterprise_orgs(user, region_code, park_code)
        org_ids = [str(org.id) for org in orgs]

        device_stats = await self._device_stats_by_org(org_ids)
        daily_stats = await self._daily_stats_by_org(org_ids, target)
        industry_by_org = await self._resolve_industry_by_org(orgs, org_ids)

        risk_distribution: dict[str, int] = {level: 0 for _, level in RISK_LEVELS}
        industry_counts: dict[str, int] = {}
        region_counts: dict[str, int] = {}

        total_devices = 0
        online_devices = 0
        offline_devices = 0

        for org in orgs:
            org_id = str(org.id)
            d_stats = device_stats.get(org_id, {"total": 0, "online": 0, "offline": 0, "alarm": 0})
            total_devices += d_stats.get("total", 0)
            online_devices += d_stats.get("online", 0)
            offline_devices += d_stats.get("offline", 0)

            score, level = self._compute_risk(d_stats, daily_stats.get(org_id, {}))
            risk_distribution[level] = risk_distribution.get(level, 0) + 1

            industry = industry_by_org.get(org_id, "other")
            industry_counts[industry] = industry_counts.get(industry, 0) + 1

            region_key = org.park_code if park_code or (org.park_code and not org.region_code) else org.region_code
            region_key = region_key or "unknown"
            region_counts[region_key] = region_counts.get(region_key, 0) + 1

        industry_distribution = []
        for industry, count in sorted(industry_counts.items(), key=lambda kv: kv[0]):
            threshold = MIN_SAMPLE_BY_INDUSTRY.get(industry, DEFAULT_MIN_SAMPLE)
            industry_distribution.append({
                "industry": industry,
                "count": count,
                "insufficient": count < threshold,
            })

        region_distribution = [
            {"region_code": key, "count": count}
            for key, count in sorted(region_counts.items(), key=lambda kv: kv[0])
        ]

        return {
            "target_date": target.isoformat(),
            "enterprise_count": len(orgs),
            "device_count": total_devices,
            "online_device_count": online_devices,
            "offline_device_count": offline_devices,
            "risk_distribution": [
                {"level": level, "count": risk_distribution.get(level, 0)}
                for _, level in RISK_LEVELS
            ],
            "industry_distribution": industry_distribution,
            "region_distribution": region_distribution,
        }

    async def get_overview_range(
        self,
        user: User,
        start_date: date,
        end_date: date,
        region_code: str | None,
        park_code: str | None,
    ) -> dict[str, Any]:
        start, end = _resolve_date_range(start_date, end_date)
        orgs = await self._fetch_enterprise_orgs(user, region_code, park_code)
        org_ids = [str(org.id) for org in orgs]

        device_stats = await self._device_stats_by_org(org_ids)
        range_stats = await self._range_stats_by_org(org_ids, start, end)
        industry_by_org = await self._resolve_industry_by_org(orgs, org_ids)

        risk_distribution: dict[str, int] = {level: 0 for _, level in RISK_LEVELS}
        industry_counts: dict[str, int] = {}
        region_counts: dict[str, int] = {}

        total_devices = 0
        online_devices = 0
        offline_devices = 0

        for org in orgs:
            org_id = str(org.id)
            d_stats = device_stats.get(org_id, {"total": 0, "online": 0, "offline": 0, "alarm": 0})
            total_devices += d_stats.get("total", 0)
            online_devices += d_stats.get("online", 0)
            offline_devices += d_stats.get("offline", 0)

            score, level = self._compute_risk(d_stats, range_stats.get(org_id, {}))
            risk_distribution[level] = risk_distribution.get(level, 0) + 1

            industry = industry_by_org.get(org_id, "other")
            industry_counts[industry] = industry_counts.get(industry, 0) + 1

            region_key = org.park_code if park_code or (org.park_code and not org.region_code) else org.region_code
            region_key = region_key or "unknown"
            region_counts[region_key] = region_counts.get(region_key, 0) + 1

        industry_distribution = []
        for industry, count in sorted(industry_counts.items(), key=lambda kv: kv[0]):
            threshold = MIN_SAMPLE_BY_INDUSTRY.get(industry, DEFAULT_MIN_SAMPLE)
            industry_distribution.append({
                "industry": industry,
                "count": count,
                "insufficient": count < threshold,
            })

        region_distribution = [
            {"region_code": key, "count": count}
            for key, count in sorted(region_counts.items(), key=lambda kv: kv[0])
        ]

        return {
            "start_date": start.isoformat(),
            "end_date": end.isoformat(),
            "enterprise_count": len(orgs),
            "device_count": total_devices,
            "online_device_count": online_devices,
            "offline_device_count": offline_devices,
            "risk_distribution": [
                {"level": level, "count": risk_distribution.get(level, 0)}
                for _, level in RISK_LEVELS
            ],
            "industry_distribution": industry_distribution,
            "region_distribution": region_distribution,
        }

    async def get_heatmap(
        self,
        user: User,
        target_date: date | None,
        resolution: int,
        region_code: str | None,
        park_code: str | None,
    ) -> dict[str, Any]:
        target = _resolve_target_date(target_date)
        orgs = await self._fetch_enterprise_orgs(user, region_code, park_code)
        org_ids = [str(org.id) for org in orgs]
        if not org_ids:
            return {"target_date": target.isoformat(), "resolution": resolution, "cells": []}

        device_stats = await self._device_stats_by_org(org_ids)
        daily_stats = await self._daily_stats_by_org(org_ids, target)

        query = select(Device.org_id, Device.latitude, Device.longitude).where(
            and_(
                Device.org_id.in_(org_ids),
                Device.latitude.isnot(None),
                Device.longitude.isnot(None),
            )
        )
        result = await self.db.execute(query)
        rows = result.all()

        cells: dict[str, dict[str, Any]] = {}
        for row in rows:
            lat = float(row.latitude)
            lng = float(row.longitude)
            cell = h3.latlng_to_cell(lat, lng, resolution)
            org_id = str(row.org_id)

            cell_info = cells.setdefault(cell, {
                "org_ids": set(),
                "device_count": 0,
                "risk_scores": [],
            })
            cell_info["device_count"] += 1
            if org_id not in cell_info["org_ids"]:
                d_stats = device_stats.get(org_id, {"total": 0, "online": 0, "offline": 0, "alarm": 0})
                score, _ = self._compute_risk(d_stats, daily_stats.get(org_id, {}))
                cell_info["org_ids"].add(org_id)
                cell_info["risk_scores"].append(score)

        payload_cells = []
        for cell, info in cells.items():
            boundary = h3.cell_to_boundary(cell)
            coords = [[lng, lat] for lat, lng in boundary]
            enterprise_count = len(info["org_ids"])
            avg_score = sum(info["risk_scores"]) / enterprise_count if enterprise_count else 0.0
            payload_cells.append({
                "h3_index": cell,
                "boundary": coords,
                "risk_level": _risk_level(avg_score),
                "risk_score": round(avg_score, 2),
                "enterprise_count": enterprise_count,
                "device_count": info["device_count"],
            })

        return {
            "target_date": target.isoformat(),
            "resolution": resolution,
            "cells": payload_cells,
        }

    async def get_trends(
        self,
        user: User,
        start_date: date | None,
        end_date: date | None,
        granularity: str,
    ) -> dict[str, Any]:
        start, end = _resolve_date_range(start_date, end_date)
        orgs = await self._fetch_enterprise_orgs(user, None, None)
        org_ids = [str(org.id) for org in orgs]
        if not org_ids:
            return {"granularity": granularity, "series": []}

        device_stats = await self._device_stats_by_org(org_ids)

        stats_query = (
            select(
                MonitoringDailyStats.stat_date,
                MonitoringDailyStats.org_id,
                func.sum(MonitoringDailyStats.data_count).label("data_count"),
                func.sum(MonitoringDailyStats.exceed_count).label("exceed_count"),
                func.sum(MonitoringDailyStats.invalid_count).label("invalid_count"),
            )
            .where(
                and_(
                    MonitoringDailyStats.stat_date >= start,
                    MonitoringDailyStats.stat_date <= end,
                    MonitoringDailyStats.org_id.in_(org_ids),
                )
            )
            .group_by(MonitoringDailyStats.stat_date, MonitoringDailyStats.org_id)
        )
        result = await self.db.execute(stats_query)
        rows = result.all()

        stats_by_date: dict[date, dict[str, dict[str, float]]] = {}
        for row in rows:
            stats_by_date.setdefault(row.stat_date, {})[str(row.org_id)] = {
                "data_count": float(row.data_count or 0),
                "exceed_count": float(row.exceed_count or 0),
                "invalid_count": float(row.invalid_count or 0),
            }

        series = []
        cursor = start
        while cursor <= end:
            distribution: dict[str, int] = {level: 0 for _, level in RISK_LEVELS}
            per_org = stats_by_date.get(cursor, {})
            for org_id in org_ids:
                d_stats = device_stats.get(org_id, {"total": 0, "online": 0, "offline": 0, "alarm": 0})
                score, level = self._compute_risk(d_stats, per_org.get(org_id, {}))
                distribution[level] = distribution.get(level, 0) + 1
            series.append({
                "date": cursor.isoformat(),
                "risk_distribution": [
                    {"level": level, "count": distribution.get(level, 0)}
                    for _, level in RISK_LEVELS
                ],
            })
            cursor += timedelta(days=1)

        if granularity == "monthly":
            monthly: dict[str, dict[str, int]] = {}
            for item in series:
                month_key = item["date"][:7]
                bucket = monthly.setdefault(month_key, {level: 0 for _, level in RISK_LEVELS})
                for entry in item["risk_distribution"]:
                    bucket[entry["level"]] += entry["count"]
            series = [
                {
                    "date": month,
                    "risk_distribution": [
                        {"level": level, "count": counts.get(level, 0)}
                        for _, level in RISK_LEVELS
                    ],
                }
                for month, counts in sorted(monthly.items())
            ]

        return {"granularity": granularity, "series": series}

    async def get_consistency(
        self,
        user: User,
        start_date: date | None,
        end_date: date | None,
    ) -> dict[str, Any]:
        start, end = _resolve_date_range(start_date, end_date, default_days=90)
        orgs = await self._fetch_enterprise_orgs(user, None, None)
        org_ids = [str(org.id) for org in orgs]
        if not org_ids:
            return {"summary": {"high": 0, "medium": 0, "low": 0}, "industry_breakdown": [], "region_breakdown": []}

        industry_by_org = await self._resolve_industry_by_org(orgs, org_ids)
        region_by_org = {str(org.id): (org.region_code or "unknown") for org in orgs}

        reports_query = (
            select(
                SelfInspectionReport.org_id,
                SelfInspectionReport.inspection_date,
                SelfInspectionData.pollutant_code,
                SelfInspectionData.value,
            )
            .join(SelfInspectionData, SelfInspectionReport.id == SelfInspectionData.report_id)
            .where(
                and_(
                    SelfInspectionReport.inspection_date >= start,
                    SelfInspectionReport.inspection_date <= end,
                    SelfInspectionReport.org_id.in_([org.id for org in orgs]),
                )
            )
        )
        report_rows = (await self.db.execute(reports_query)).all()

        stats_query = (
            select(
                MonitoringDailyStats.org_id,
                MonitoringDailyStats.stat_date,
                MonitoringDailyStats.pollutant_code,
                MonitoringDailyStats.avg_value,
            )
            .where(
                and_(
                    MonitoringDailyStats.stat_date >= start,
                    MonitoringDailyStats.stat_date <= end,
                    MonitoringDailyStats.org_id.in_(org_ids),
                )
            )
        )
        stats_rows = (await self.db.execute(stats_query)).all()

        avg_lookup: dict[tuple[str, date, str], float] = {}
        for row in stats_rows:
            if row.avg_value is None:
                continue
            avg_lookup[(str(row.org_id), row.stat_date, row.pollutant_code)] = float(row.avg_value)

        per_org: dict[str, dict[str, int]] = {}
        for row in report_rows:
            org_id = str(row.org_id)
            value = _parse_numeric_value(row.value)
            if value is None:
                continue
            key = (org_id, row.inspection_date, row.pollutant_code)
            if key not in avg_lookup:
                continue
            avg_value = avg_lookup[key]
            if avg_value <= 0:
                continue
            diff_ratio = abs(value - avg_value) / avg_value
            consistent = diff_ratio <= 0.3
            stats = per_org.setdefault(org_id, {"total": 0, "consistent": 0})
            stats["total"] += 1
            if consistent:
                stats["consistent"] += 1

        summary = {"high": 0, "medium": 0, "low": 0}
        industry_breakdown: dict[str, dict[str, int]] = {}
        region_breakdown: dict[str, dict[str, int]] = {}

        for org_id, stats in per_org.items():
            total = stats["total"]
            if total <= 0:
                continue
            score = stats["consistent"] / total
            if score >= 0.8:
                bucket = "high"
            elif score >= 0.5:
                bucket = "medium"
            else:
                bucket = "low"
            summary[bucket] += 1

            industry = industry_by_org.get(org_id, "other")
            industry_breakdown.setdefault(industry, {"high": 0, "medium": 0, "low": 0})
            industry_breakdown[industry][bucket] += 1

            region = region_by_org.get(org_id, "unknown")
            region_breakdown.setdefault(region, {"high": 0, "medium": 0, "low": 0})
            region_breakdown[region][bucket] += 1

        industry_payload = [
            {"industry": key, **counts}
            for key, counts in sorted(industry_breakdown.items())
        ]
        region_payload = [
            {"region_code": key, **counts}
            for key, counts in sorted(region_breakdown.items())
        ]

        return {
            "summary": summary,
            "industry_breakdown": industry_payload,
            "region_breakdown": region_payload,
        }

    def _usage_key(self, period: str) -> date:
        now = datetime.now(ZoneInfo("Asia/Shanghai")).date()
        if period == "monthly":
            return date(now.year, now.month, 1)
        return now

    async def _consume_brief_quota(
        self,
        user: User,
        period: str,
        target_label: str | None,
    ) -> dict[str, int]:
        limit = BRIEF_LIMITS.get(period, 0)
        if limit <= 0:
            return {"limit": 0, "remaining": 0}

        usage_date = self._usage_key(period)
        result = await self.db.execute(
            select(func.count(RegulatorBriefUsage.id)).where(
                RegulatorBriefUsage.user_id == user.id,
                RegulatorBriefUsage.period == period,
                RegulatorBriefUsage.usage_date == usage_date,
            )
        )
        used = int(result.scalar() or 0)
        if used >= limit:
            raise RateLimitError(period=period, limit=limit)

        self.db.add(RegulatorBriefUsage(
            user_id=user.id,
            period=period,
            usage_date=usage_date,
            target_label=target_label,
        ))
        remaining = max(limit - used - 1, 0)
        return {"limit": limit, "remaining": remaining}

    async def _summarize_period(
        self,
        user: User,
        start: date,
        end: date,
        region_code: str | None,
        park_code: str | None,
        use_range: bool,
    ) -> dict[str, Any]:
        orgs = await self._fetch_enterprise_orgs(user, region_code, park_code)
        org_ids = [str(org.id) for org in orgs]
        device_stats = await self._device_stats_by_org(org_ids)
        if use_range:
            stats_by_org = await self._range_stats_by_org(org_ids, start, end)
        else:
            stats_by_org = await self._daily_stats_by_org(org_ids, start)

        industry_by_org = await self._resolve_industry_by_org(orgs, org_ids)

        risk_distribution: dict[str, int] = {level: 0 for _, level in RISK_LEVELS}
        industry_counts: dict[str, int] = {}
        region_counts: dict[str, int] = {}
        region_names: dict[str, str] = {}
        industry_groups: dict[str, dict[str, Any]] = {}
        region_groups: dict[str, dict[str, Any]] = {}

        total_devices = 0
        online_devices = 0
        offline_devices = 0
        alarm_devices = 0
        total_data = 0.0
        total_exceed = 0.0
        total_invalid = 0.0
        risk_scores: list[float] = []

        for org in orgs:
            org_id = str(org.id)
            d_stats = device_stats.get(org_id, {"total": 0, "online": 0, "offline": 0, "alarm": 0})
            s_stats = stats_by_org.get(org_id, {"data_count": 0, "exceed_count": 0, "invalid_count": 0})

            total_devices += d_stats.get("total", 0)
            online_devices += d_stats.get("online", 0)
            offline_devices += d_stats.get("offline", 0)
            alarm_devices += d_stats.get("alarm", 0)
            total_data += float(s_stats.get("data_count", 0))
            total_exceed += float(s_stats.get("exceed_count", 0))
            total_invalid += float(s_stats.get("invalid_count", 0))

            score, level = self._compute_risk(d_stats, s_stats)
            risk_scores.append(score)
            risk_distribution[level] = risk_distribution.get(level, 0) + 1

            industry = industry_by_org.get(org_id, "other")
            industry_counts[industry] = industry_counts.get(industry, 0) + 1
            industry_groups.setdefault(industry, {"scores": [], "count": 0})
            industry_groups[industry]["scores"].append(score)
            industry_groups[industry]["count"] += 1

            region_key = org.park_code if park_code or (org.park_code and not org.region_code) else org.region_code
            region_key = region_key or "unknown"
            region_counts[region_key] = region_counts.get(region_key, 0) + 1
            region_groups.setdefault(region_key, {"scores": [], "count": 0})
            region_groups[region_key]["scores"].append(score)
            region_groups[region_key]["count"] += 1

            if region_key not in region_names:
                if region_key == org.park_code and org.park_name:
                    region_names[region_key] = org.park_name
                elif region_key == org.region_code and org.region_name:
                    region_names[region_key] = org.region_name
                else:
                    region_names[region_key] = region_key

        industry_distribution = []
        for industry, count in sorted(industry_counts.items(), key=lambda kv: kv[0]):
            threshold = MIN_SAMPLE_BY_INDUSTRY.get(industry, DEFAULT_MIN_SAMPLE)
            industry_distribution.append({
                "industry": industry,
                "count": count,
                "insufficient": count < threshold,
            })
            industry_groups[industry]["threshold"] = threshold

        region_distribution = [
            {"region_code": key, "count": count}
            for key, count in sorted(region_counts.items(), key=lambda kv: kv[0])
        ]

        average_score = sum(risk_scores) / len(risk_scores) if risk_scores else 0.0
        return {
            "enterprise_count": len(orgs),
            "device_count": total_devices,
            "online_device_count": online_devices,
            "offline_device_count": offline_devices,
            "alarm_device_count": alarm_devices,
            "data_count": total_data,
            "exceed_count": total_exceed,
            "invalid_count": total_invalid,
            "average_risk_score": round(average_score, 2),
            "overall_risk_level": _risk_level(average_score),
            "risk_distribution": [
                {"level": level, "count": risk_distribution.get(level, 0)}
                for _, level in RISK_LEVELS
            ],
            "industry_distribution": industry_distribution,
            "region_distribution": region_distribution,
            "industry_groups": industry_groups,
            "region_groups": region_groups,
            "region_names": region_names,
        }

    def _rank_groups(
        self,
        groups: dict[str, dict[str, Any]],
        total_enterprises: int,
        top_n: int,
        *,
        names: dict[str, str] | None = None,
        filter_insufficient: bool = False,
    ) -> list[dict[str, Any]]:
        items: list[dict[str, Any]] = []
        for key, info in groups.items():
            count = int(info.get("count", 0))
            scores = info.get("scores", [])
            avg_score = sum(scores) / len(scores) if scores else 0.0
            threshold = info.get("threshold", DEFAULT_MIN_SAMPLE)
            insufficient = count < threshold
            if filter_insufficient and insufficient:
                continue
            items.append({
                "code": key,
                "name": (names or {}).get(key, key),
                "risk_score": round(avg_score, 2),
                "risk_level": _risk_level(avg_score),
                "enterprise_count": count,
                "share": round(_safe_divide(count, total_enterprises) * 100, 2),
                "insufficient": insufficient,
            })

        items.sort(key=lambda item: item["risk_score"], reverse=True)
        return items[:top_n]

    async def get_ai_brief(
        self,
        user: User,
        period: str,
        target_date: date | None,
        year: int | None,
        month: int | None,
        region_code: str | None,
        park_code: str | None,
    ) -> dict[str, Any]:
        period = period.lower()
        if period not in {"daily", "monthly"}:
            raise ValueError("invalid period")

        if period == "daily":
            target = _resolve_target_date(target_date)
            start = target
            end = target
            prev_start = target - timedelta(days=1)
            prev_end = prev_start
            label = target.isoformat()
            use_range = False
        else:
            now = datetime.now(ZoneInfo("Asia/Shanghai")).date()
            year = year or now.year
            month = month or now.month
            start, end = _resolve_month_range(year, month)
            end = _resolve_target_date(end)
            if start > end:
                start = end
            prev_month = start - timedelta(days=1)
            prev_start, prev_end = _resolve_month_range(prev_month.year, prev_month.month)
            prev_end = _resolve_target_date(prev_end)
            if prev_start > prev_end:
                prev_start = prev_end
            label = start.strftime("%Y-%m")
            use_range = True

        quota = await self._consume_brief_quota(user, period, label)

        current = await self._summarize_period(user, start, end, region_code, park_code, use_range)
        previous = await self._summarize_period(user, prev_start, prev_end, region_code, park_code, use_range)

        avg_score = current["average_risk_score"]
        prev_score = previous["average_risk_score"]
        delta = round(avg_score - prev_score, 2)

        if previous["enterprise_count"] <= 0:
            trend_text = "暂无可比数据"
        elif delta >= 1.0:
            trend_text = f"较上期上升{delta:.1f}分"
        elif delta <= -1.0:
            trend_text = f"较上期下降{abs(delta):.1f}分"
        else:
            trend_text = "较上期基本持平"

        total_devices = current["device_count"]
        data_count = current["data_count"]
        exceed_rate = _safe_divide(current["exceed_count"], data_count)
        invalid_rate = _safe_divide(current["invalid_count"], data_count)
        offline_rate = _safe_divide(current["offline_device_count"], total_devices)
        alarm_rate = _safe_divide(current["alarm_device_count"], total_devices)
        online_rate = _safe_divide(current["online_device_count"], total_devices)

        driver_candidates = [
            ("超标率", exceed_rate),
            ("无效率", invalid_rate),
            ("离线率", offline_rate),
            ("报警率", alarm_rate),
        ]
        driver_candidates.sort(key=lambda item: item[1], reverse=True)
        drivers = [name for name, _ in driver_candidates[:2] if _ > 0]
        driver_text = "、".join(drivers) if drivers else "暂无明显异常指标"

        overall_summary = (
            f"本期综合风险等级为{current['overall_risk_level']}，{trend_text}，"
            f"综合风险分为{avg_score:.1f}。主要风险因子为{driver_text}。"
        )

        top_regions = self._rank_groups(
            current["region_groups"],
            current["enterprise_count"],
            BRIEF_TOP_N,
            names=current["region_names"],
        )

        top_industries = self._rank_groups(
            current["industry_groups"],
            current["enterprise_count"],
            BRIEF_TOP_N,
            filter_insufficient=True,
        )

        insufficient_industries = [
            item["industry"]
            for item in current["industry_distribution"]
            if item["insufficient"]
        ]

        highlights: list[str] = []
        if top_regions:
            names = "、".join([item["name"] for item in top_regions])
            highlights.append(f"风险较高区域集中于：{names}。")
        if top_industries:
            names = "、".join([item["code"] for item in top_industries])
            highlights.append(f"高风险行业集中于：{names}。")
        if insufficient_industries:
            names = "、".join(insufficient_industries)
            highlights.append(f"样本不足行业不纳入排名：{names}。")

        suggestion_map = {
            "超标率": "建议对高风险行业与区域开展超标排查与复测。",
            "无效率": "建议督促企业提升数据有效性，规范自测与数据采集流程。",
            "离线率": "建议对离线设备较多区域开展在线核查与维护。",
            "报警率": "建议加强告警处置与闭环管理，降低异常积压。",
        }
        suggestions = [suggestion_map[d] for d in drivers if d in suggestion_map]
        if not suggestions:
            suggestions = ["建议持续开展重点区域巡查，强化风险预警与执法协同。"]

        consistency = await self.get_consistency(user, start, end)

        return {
            "period": period,
            "period_label": label,
            "start_date": start.isoformat(),
            "end_date": end.isoformat(),
            "scope": {"region_code": region_code, "park_code": park_code},
            "summary_text": overall_summary,
            "highlights": highlights,
            "suggestions": suggestions,
            "quota": quota,
            "overview": {
                "enterprise_count": current["enterprise_count"],
                "device_count": current["device_count"],
                "online_rate": round(online_rate * 100, 2),
                "offline_rate": round(offline_rate * 100, 2),
                "alarm_rate": round(alarm_rate * 100, 2),
                "exceed_rate": round(exceed_rate * 100, 2),
                "invalid_rate": round(invalid_rate * 100, 2),
                "risk_level": current["overall_risk_level"],
                "risk_score": avg_score,
                "risk_trend": trend_text,
            },
            "top_regions": top_regions,
            "top_industries": top_industries,
            "industry_distribution": current["industry_distribution"],
            "consistency": consistency,
            "data_note": "本简报仅展示聚合统计结果，不含企业级明细数据。",
        }

    async def generate_brief_excel(self, payload: dict[str, Any]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "监管简报"

        header_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_text = Font(bold=True, color="FFFFFF")

        ws.merge_cells("A1:F1")
        ws["A1"] = "AI监管简报"
        ws["A1"].font = header_font
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A3"] = "周期"
        ws["B3"] = payload["period_label"]
        ws["A4"] = "统计范围"
        ws["B4"] = payload["scope"].get("park_code") or payload["scope"].get("region_code") or "全域"
        ws["A5"] = "统计区间"
        ws["B5"] = f"{payload['start_date']} 至 {payload['end_date']}"

        ws["A7"] = "总体态势"
        ws["A7"].font = section_font
        ws.merge_cells("A8:F9")
        ws["A8"] = payload["summary_text"]

        overview = payload["overview"]
        ws["A11"] = "运行质量"
        ws["A11"].font = section_font
        metrics = [
            ("企业数", overview["enterprise_count"]),
            ("设备数", overview["device_count"]),
            ("在线率", f"{overview['online_rate']}%"),
            ("离线率", f"{overview['offline_rate']}%"),
            ("超标率", f"{overview['exceed_rate']}%"),
            ("无效率", f"{overview['invalid_rate']}%"),
            ("报警率", f"{overview['alarm_rate']}%"),
        ]
        row = 12
        for label, value in metrics:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1

        row += 1
        ws[f"A{row}"] = "区域态势TOP5"
        ws[f"A{row}"].font = section_font
        row += 1
        headers = ["区域", "风险等级", "风险分", "企业数", "占比"]
        for idx, header in enumerate(headers, 0):
            cell = ws.cell(row=row, column=1 + idx, value=header)
            cell.font = header_text
            cell.fill = header_fill
            cell.border = thin_border
        row += 1
        for item in payload["top_regions"]:
            ws.cell(row=row, column=1, value=item["name"]).border = thin_border
            ws.cell(row=row, column=2, value=item["risk_level"]).border = thin_border
            ws.cell(row=row, column=3, value=item["risk_score"]).border = thin_border
            ws.cell(row=row, column=4, value=item["enterprise_count"]).border = thin_border
            ws.cell(row=row, column=5, value=f"{item['share']}%").border = thin_border
            row += 1

        row += 1
        ws[f"A{row}"] = "行业态势TOP5"
        ws[f"A{row}"].font = section_font
        row += 1
        headers = ["行业", "风险等级", "风险分", "企业数", "占比"]
        for idx, header in enumerate(headers, 0):
            cell = ws.cell(row=row, column=1 + idx, value=header)
            cell.font = header_text
            cell.fill = header_fill
            cell.border = thin_border
        row += 1
        for item in payload["top_industries"]:
            ws.cell(row=row, column=1, value=item["code"]).border = thin_border
            ws.cell(row=row, column=2, value=item["risk_level"]).border = thin_border
            ws.cell(row=row, column=3, value=item["risk_score"]).border = thin_border
            ws.cell(row=row, column=4, value=item["enterprise_count"]).border = thin_border
            ws.cell(row=row, column=5, value=f"{item['share']}%").border = thin_border
            row += 1

        row += 1
        ws[f"A{row}"] = "一致性评估"
        ws[f"A{row}"].font = section_font
        row += 1
        summary = payload["consistency"]["summary"]
        ws.cell(row=row, column=1, value="高一致").border = thin_border
        ws.cell(row=row, column=2, value=summary.get("high", 0)).border = thin_border
        row += 1
        ws.cell(row=row, column=1, value="中一致").border = thin_border
        ws.cell(row=row, column=2, value=summary.get("medium", 0)).border = thin_border
        row += 1
        ws.cell(row=row, column=1, value="低一致").border = thin_border
        ws.cell(row=row, column=2, value=summary.get("low", 0)).border = thin_border
        row += 2

        ws[f"A{row}"] = "监管建议"
        ws[f"A{row}"].font = section_font
        row += 1
        for suggestion in payload["suggestions"]:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row=row, column=1, value=suggestion)
            row += 1

        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=payload["data_note"])

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 24

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    async def generate_brief_pdf(self, payload: dict[str, Any]) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=18 * mm,
            leftMargin=18 * mm,
            topMargin=16 * mm,
            bottomMargin=16 * mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "BriefTitle",
            parent=styles["Heading1"],
            fontName=CHINESE_FONT,
            fontSize=16,
            alignment=1,
            spaceAfter=12,
        )
        section_style = ParagraphStyle(
            "BriefSection",
            parent=styles["Heading2"],
            fontName=CHINESE_FONT,
            fontSize=12,
            spaceAfter=6,
        )
        normal_style = ParagraphStyle(
            "BriefNormal",
            parent=styles["Normal"],
            fontName=CHINESE_FONT,
            fontSize=9,
            spaceAfter=4,
        )

        elements: list[Any] = []
        elements.append(Paragraph("AI监管简报", title_style))
        elements.append(Paragraph(f"周期：{payload['period_label']}", normal_style))
        scope_label = payload["scope"].get("park_code") or payload["scope"].get("region_code") or "全域"
        elements.append(Paragraph(f"统计范围：{scope_label}", normal_style))
        elements.append(Paragraph(f"统计区间：{payload['start_date']} 至 {payload['end_date']}", normal_style))
        elements.append(Spacer(1, 8))

        elements.append(Paragraph("总体态势", section_style))
        elements.append(Paragraph(payload["summary_text"], normal_style))
        elements.append(Spacer(1, 6))

        overview = payload["overview"]
        metrics = [
            ["企业数", str(overview["enterprise_count"])],
            ["设备数", str(overview["device_count"])],
            ["在线率", f"{overview['online_rate']}%"],
            ["离线率", f"{overview['offline_rate']}%"],
            ["超标率", f"{overview['exceed_rate']}%"],
            ["无效率", f"{overview['invalid_rate']}%"],
            ["报警率", f"{overview['alarm_rate']}%"],
        ]
        table = Table([["指标", "数值"]] + metrics, colWidths=[90, 90])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, -1), CHINESE_FONT),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 8))

        elements.append(Paragraph("区域态势TOP5", section_style))
        region_rows = [["区域", "风险等级", "风险分", "企业数"]]
        for item in payload["top_regions"]:
            region_rows.append([item["name"], item["risk_level"], str(item["risk_score"]), str(item["enterprise_count"])])
        region_table = Table(region_rows, colWidths=[110, 60, 60, 50])
        region_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, -1), CHINESE_FONT),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(region_table)
        elements.append(Spacer(1, 8))

        elements.append(Paragraph("行业态势TOP5", section_style))
        industry_rows = [["行业", "风险等级", "风险分", "企业数"]]
        for item in payload["top_industries"]:
            industry_rows.append([item["code"], item["risk_level"], str(item["risk_score"]), str(item["enterprise_count"])])
        industry_table = Table(industry_rows, colWidths=[110, 60, 60, 50])
        industry_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, -1), CHINESE_FONT),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(industry_table)
        elements.append(Spacer(1, 8))

        elements.append(Paragraph("一致性评估", section_style))
        summary = payload["consistency"]["summary"]
        consistency_table = Table(
            [["高一致", summary.get("high", 0)], ["中一致", summary.get("medium", 0)], ["低一致", summary.get("low", 0)]],
            colWidths=[90, 60]
        )
        consistency_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, -1), CHINESE_FONT),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(consistency_table)
        elements.append(Spacer(1, 8))

        elements.append(Paragraph("监管建议", section_style))
        for suggestion in payload["suggestions"]:
            elements.append(Paragraph(suggestion, normal_style))

        elements.append(Spacer(1, 6))
        elements.append(Paragraph(payload["data_note"], normal_style))

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    async def build_report_payload(
        self,
        user: User,
        report_type: str,
        target_date: date | None,
        year: int | None,
        month: int | None,
        region_code: str | None,
        park_code: str | None,
    ) -> dict[str, Any]:
        report_type = report_type.lower()
        if report_type not in {"daily", "monthly"}:
            raise ValueError("invalid report_type")

        if report_type == "daily":
            target = _resolve_target_date(target_date)
            overview = await self.get_overview(user, target, region_code, park_code)
            period_label = target.isoformat()
            return {
                "report_type": "daily",
                "period_label": period_label,
                "start_date": target.isoformat(),
                "end_date": target.isoformat(),
                "overview": overview,
            }

        if year is None or month is None:
            now = datetime.now(ZoneInfo("Asia/Shanghai")).date()
            year = now.year
            month = now.month

        start, end = _resolve_month_range(year, month)
        end = _resolve_target_date(end)
        if start > end:
            start = end
        overview = await self.get_overview_range(user, start, end, region_code, park_code)
        period_label = f"{start.strftime('%Y-%m')}"
        return {
            "report_type": "monthly",
            "period_label": period_label,
            "start_date": start.isoformat(),
            "end_date": end.isoformat(),
            "overview": overview,
        }

    async def generate_excel_report(self, payload: dict[str, Any]) -> bytes:
        overview = payload["overview"]

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        header_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_text = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        ws.merge_cells("A1:D1")
        ws["A1"] = "Regulator Summary Report"
        ws["A1"].font = header_font
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A3"] = "Report Type"
        ws["B3"] = payload["report_type"]
        ws["A4"] = "Period"
        ws["B4"] = payload["period_label"]
        ws["A5"] = "Start Date"
        ws["B5"] = payload["start_date"]
        ws["A6"] = "End Date"
        ws["B6"] = payload["end_date"]

        summary_rows = [
            ("Enterprise Count", overview["enterprise_count"]),
            ("Device Count", overview["device_count"]),
            ("Online Devices", overview["online_device_count"]),
            ("Offline Devices", overview["offline_device_count"]),
        ]

        ws["A8"] = "Summary"
        ws["A8"].font = section_font
        row = 9
        for label, value in summary_rows:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1

        row += 1
        ws[f"A{row}"] = "Risk Distribution"
        ws[f"A{row}"].font = section_font
        row += 1
        ws[f"A{row}"] = "Level"
        ws[f"B{row}"] = "Count"
        ws[f"A{row}"].font = header_text
        ws[f"B{row}"].font = header_text
        ws[f"A{row}"].fill = header_fill
        ws[f"B{row}"].fill = header_fill
        ws[f"A{row}"].border = thin_border
        ws[f"B{row}"].border = thin_border
        row += 1
        for entry in overview["risk_distribution"]:
            ws[f"A{row}"] = entry["level"]
            ws[f"B{row}"] = entry["count"]
            ws[f"A{row}"].border = thin_border
            ws[f"B{row}"].border = thin_border
            row += 1

        row += 1
        ws[f"A{row}"] = "Industry Distribution"
        ws[f"A{row}"].font = section_font
        row += 1
        ws[f"A{row}"] = "Industry"
        ws[f"B{row}"] = "Count"
        ws[f"C{row}"] = "Insufficient"
        for col in ("A", "B", "C"):
            ws[f"{col}{row}"].font = header_text
            ws[f"{col}{row}"].fill = header_fill
            ws[f"{col}{row}"].border = thin_border
        row += 1
        for entry in overview["industry_distribution"]:
            ws[f"A{row}"] = entry["industry"]
            ws[f"B{row}"] = entry["count"]
            ws[f"C{row}"] = "yes" if entry["insufficient"] else "no"
            for col in ("A", "B", "C"):
                ws[f"{col}{row}"].border = thin_border
            row += 1

        row += 1
        ws[f"A{row}"] = "Region Distribution"
        ws[f"A{row}"].font = section_font
        row += 1
        ws[f"A{row}"] = "Region"
        ws[f"B{row}"] = "Count"
        for col in ("A", "B"):
            ws[f"{col}{row}"].font = header_text
            ws[f"{col}{row}"].fill = header_fill
            ws[f"{col}{row}"].border = thin_border
        row += 1
        for entry in overview["region_distribution"]:
            ws[f"A{row}"] = entry["region_code"]
            ws[f"B{row}"] = entry["count"]
            ws[f"A{row}"].border = thin_border
            ws[f"B{row}"].border = thin_border
            row += 1

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    async def generate_pdf_report(self, payload: dict[str, Any]) -> bytes:
        overview = payload["overview"]

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=18 * mm,
            leftMargin=18 * mm,
            topMargin=16 * mm,
            bottomMargin=16 * mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "RegulatorTitle",
            parent=styles["Heading1"],
            fontSize=16,
            alignment=1,
            spaceAfter=12,
        )
        section_style = ParagraphStyle(
            "RegulatorSection",
            parent=styles["Heading2"],
            fontSize=12,
            spaceAfter=6,
        )
        normal_style = ParagraphStyle(
            "RegulatorNormal",
            parent=styles["Normal"],
            fontSize=9,
            spaceAfter=4,
        )

        elements: list[Any] = []
        elements.append(Paragraph("Regulator Summary Report", title_style))
        elements.append(Paragraph(f"Report Type: {payload['report_type']}", normal_style))
        elements.append(Paragraph(f"Period: {payload['period_label']}", normal_style))
        elements.append(Paragraph(f"Start Date: {payload['start_date']}", normal_style))
        elements.append(Paragraph(f"End Date: {payload['end_date']}", normal_style))
        elements.append(Spacer(1, 8))

        summary_data = [
            ["Metric", "Value"],
            ["Enterprise Count", str(overview["enterprise_count"])],
            ["Device Count", str(overview["device_count"])],
            ["Online Devices", str(overview["online_device_count"])],
            ["Offline Devices", str(overview["offline_device_count"])],
        ]
        summary_table = Table(summary_data, colWidths=[120, 120])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(Paragraph("Summary", section_style))
        elements.append(summary_table)
        elements.append(Spacer(1, 10))

        risk_data = [["Level", "Count"]]
        for entry in overview["risk_distribution"]:
            risk_data.append([entry["level"], str(entry["count"])])
        risk_table = Table(risk_data, colWidths=[80, 80])
        risk_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(Paragraph("Risk Distribution", section_style))
        elements.append(risk_table)
        elements.append(Spacer(1, 10))

        industry_data = [["Industry", "Count", "Insufficient"]]
        for entry in overview["industry_distribution"]:
            industry_data.append([
                entry["industry"],
                str(entry["count"]),
                "yes" if entry["insufficient"] else "no",
            ])
        industry_table = Table(industry_data, colWidths=[160, 60, 80])
        industry_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(Paragraph("Industry Distribution", section_style))
        elements.append(industry_table)
        elements.append(Spacer(1, 10))

        region_data = [["Region", "Count"]]
        for entry in overview["region_distribution"]:
            region_data.append([entry["region_code"], str(entry["count"])])
        region_table = Table(region_data, colWidths=[160, 60])
        region_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(Paragraph("Region Distribution", section_style))
        elements.append(region_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
